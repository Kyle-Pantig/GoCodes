"""
PDF generation utility for automated reports
"""
import io
from typing import Dict, Any, List, Optional
from datetime import datetime, timezone, timedelta
import logging

logger = logging.getLogger(__name__)

# Use same timezone as report_schedule (UTC+8 for Philippines)
TIMEZONE_OFFSET_HOURS = 8
LOCAL_TIMEZONE = timezone(timedelta(hours=TIMEZONE_OFFSET_HOURS))

try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    FPDF = None
    PDF_AVAILABLE = False


class ReportPDF(FPDF):
    """Custom PDF class for report generation"""
    
    def __init__(self, report_name: str, report_type: str):
        # A4 Landscape: 297mm x 210mm
        super().__init__(orientation='L', format='A4')
        self.report_name = report_name
        self.report_type = report_type
        # Set explicit margins: left, top, right
        self.set_margins(left=10, top=10, right=10)
        self.set_auto_page_break(auto=True, margin=15)
        self._header_printed = False
        # Store generated time in local timezone
        now_utc = datetime.now(timezone.utc)
        now_local = now_utc.astimezone(LOCAL_TIMEZONE)
        self._generated_time = now_local.strftime("%Y-%m-%d %H:%M:%S")
        
    def header(self):
        """PDF Header - only show title on first page"""
        if not self._header_printed:
            # Full header on first page only
            self.set_font('Helvetica', 'B', 16)
            self.set_text_color(102, 126, 234)  # Purple color from email template
            self.cell(0, 10, self.report_name, new_x='LMARGIN', new_y='NEXT', align='C')
            self.set_font('Helvetica', '', 10)
            self.set_text_color(128, 128, 128)
            self.cell(0, 5, f'Generated: {self._generated_time}', new_x='LMARGIN', new_y='NEXT', align='C')
            self.ln(10)
            self._header_printed = True
        else:
            # Minimal header on subsequent pages - just some spacing
            self.ln(5)
        
    def footer(self):
        """PDF Footer"""
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Page {self.page_no()}', align='C')

    def add_section_title(self, title: str):
        """Add a section title"""
        self.set_font('Helvetica', 'B', 12)
        self.set_text_color(51, 51, 51)
        self.set_fill_color(240, 240, 240)
        self.cell(0, 8, title, new_x='LMARGIN', new_y='NEXT', fill=True)
        self.ln(2)
        
    def add_summary_row(self, label: str, value: str):
        """Add a summary row with label and value"""
        self.set_font('Helvetica', '', 10)
        self.set_text_color(51, 51, 51)
        self.cell(90, 6, label, new_x='RIGHT')
        self.set_font('Helvetica', 'B', 10)
        self.cell(0, 6, str(value), new_x='LMARGIN', new_y='NEXT')

    def _get_text_height(self, text: str, width: float, font_size: int = 7) -> float:
        """Calculate required height for text in a cell with wrapping"""
        self.set_font('Helvetica', '', font_size)
        # Get string width and calculate lines needed
        if not text:
            return 6
        
        # Calculate available width (minus padding)
        available_width = width - 2
        if available_width <= 0:
            return 6
            
        text_width = self.get_string_width(text)
        if text_width <= available_width:
            return 6
            
        # Estimate lines needed more accurately
        # Split by words and calculate actual wrapping
        words = text.split()
        lines = 1
        current_line_width = 0
        
        for word in words:
            word_width = self.get_string_width(word + ' ')
            if current_line_width + word_width > available_width:
                lines += 1
                current_line_width = word_width
            else:
                current_line_width += word_width
        
        # Each line is approximately 4 units high, plus padding
        return max(6, lines * 4 + 2)

    def _calculate_row_height(self, row: List[str], col_widths: List[int]) -> float:
        """Calculate the maximum height needed for a row based on text wrapping"""
        max_height = 6  # Minimum height
        for i, cell in enumerate(row):
            w = col_widths[i] if i < len(col_widths) else col_widths[-1]
            text = str(cell) if cell is not None else ''
            height = self._get_text_height(text, w)
            max_height = max(max_height, height)
        return max_height
        
    def add_table(self, headers: List[str], rows: List[List[str]], col_widths: Optional[List[int]] = None):
        """Add a table to the PDF with text wrapping support"""
        if not headers or not rows:
            return
            
        # Calculate column widths based on content if not provided
        # Account for left margin and right margin (r_margin defaults to l_margin if not set separately)
        page_width = self.w - self.l_margin - self.r_margin
        num_cols = len(headers)
        
        if col_widths is None:
            # Smart column width calculation based on header names and content
            col_widths = self._calculate_smart_widths(headers, rows, page_width)
        else:
            # Ensure widths fit page
            total = sum(col_widths)
            if total > page_width:
                ratio = page_width / total
                col_widths = [int(w * ratio) for w in col_widths]
        
        # Header row
        self._draw_header_row(headers, col_widths)
        
        # Data rows
        self.set_font('Helvetica', '', 7)
        self.set_text_color(51, 51, 51)
        row_fill = False
        
        # Calculate page bottom margin (footer space + safety buffer)
        page_bottom = self.h - 25  # 25mm from bottom to account for footer and buffer
        
        for row in rows:
            # Calculate row height based on content
            row_height = self._calculate_row_height(row, col_widths)
            
            # Add extra buffer for safety
            row_height_with_buffer = row_height + 2
            
            # Check if we need a new page BEFORE drawing
            if self.get_y() + row_height_with_buffer > page_bottom:
                self.add_page()
                self._draw_header_row(headers, col_widths)
                self.set_font('Helvetica', '', 7)
                self.set_text_color(51, 51, 51)
                
            if row_fill:
                self.set_fill_color(248, 248, 248)
            else:
                self.set_fill_color(255, 255, 255)
            
            # Temporarily disable auto page break during row drawing
            auto_pb = self.auto_page_break
            self.set_auto_page_break(False)
            
            # Draw cells with text wrapping support
            start_x = self.l_margin
            start_y = self.get_y()
            
            # First pass: draw all cell backgrounds and borders
            for i, cell in enumerate(row):
                w = col_widths[i] if i < len(col_widths) else col_widths[-1]
                cell_x = start_x + sum(col_widths[:i])
                self.set_xy(cell_x, start_y)
                self.cell(w, row_height, '', border=1, fill=True)
            
            # Second pass: draw text with wrapping
            self.set_font('Helvetica', '', 7)
            for i, cell in enumerate(row):
                w = col_widths[i] if i < len(col_widths) else col_widths[-1]
                text = str(cell) if cell is not None else ''
                
                cell_x = start_x + sum(col_widths[:i])
                self.set_xy(cell_x + 1, start_y + 1)
                # Use multi_cell for text wrapping, but limit to cell height
                self.multi_cell(w - 2, 4, text, border=0, align='L')
            
            # Move to next row
            self.set_xy(start_x, start_y + row_height)
            
            # Restore auto page break
            self.set_auto_page_break(auto_pb, margin=15)
            row_fill = not row_fill

    def _draw_header_row(self, headers: List[str], col_widths: List[int]):
        """Draw the header row with text wrapping support"""
        # Temporarily disable auto page break to prevent unwanted breaks during row drawing
        auto_pb = self.auto_page_break
        self.set_auto_page_break(False)
        
        self.set_font('Helvetica', 'B', 7)
        self.set_fill_color(102, 126, 234)
        self.set_text_color(255, 255, 255)
        
        # Calculate header height based on longest header
        header_height = 8  # Minimum height
        for i, header in enumerate(headers):
            w = col_widths[i] if i < len(col_widths) else col_widths[-1]
            text = str(header) if header else ''
            height = self._get_text_height(text, w, font_size=7)
            header_height = max(header_height, height)
        
        start_x = self.l_margin
        start_y = self.get_y()
        
        # First pass: draw all cell backgrounds and borders
        for i, header in enumerate(headers):
            w = col_widths[i] if i < len(col_widths) else col_widths[-1]
            cell_x = start_x + sum(col_widths[:i])
            self.set_xy(cell_x, start_y)
            self.cell(w, header_height, '', border=1, fill=True)
        
        # Second pass: draw text with wrapping
        self.set_font('Helvetica', 'B', 7)
        for i, header in enumerate(headers):
            w = col_widths[i] if i < len(col_widths) else col_widths[-1]
            text = str(header) if header else ''
            
            cell_x = start_x + sum(col_widths[:i])
            self.set_xy(cell_x + 1, start_y + 1)
            self.multi_cell(w - 2, 4, text, border=0, align='C')
        
        # Move to next line after header
        self.set_xy(start_x, start_y + header_height)
        
        # Restore auto page break
        self.set_auto_page_break(auto_pb, margin=15)
        
    def _calculate_smart_widths(self, headers: List[str], rows: List[List[str]], page_width: float) -> List[int]:
        """Calculate smart column widths based on content"""
        num_cols = len(headers)
        
        # For tables with many columns (like inventory with 16 columns), use compact widths
        if num_cols >= 14:
            # Very compact mode for many columns
            min_width = 12
            max_width = 35
        elif num_cols >= 10:
            # Compact mode
            min_width = 14
            max_width = 45
        else:
            # Normal mode
            min_width = 15
            max_width = 60
        
        # Calculate widths based on header length and sample content
        widths = []
        for i, header in enumerate(headers):
            header_lower = header.lower()
            
            # Assign widths based on column type - use compact values for many columns
            if 'item code' in header_lower or 'itemcode' in header_lower:
                widths.append(22 if num_cols < 14 else 18)
            elif 'id' in header_lower and 'tag' not in header_lower:
                widths.append(18 if num_cols < 14 else 15)
            elif 'tag' in header_lower:
                widths.append(30 if num_cols < 14 else 22)
            elif 'stock' in header_lower:
                widths.append(14)  # Narrow column for stock numbers
            elif 'unit' == header_lower or header_lower == 'unit':
                widths.append(14)  # Narrow for unit
            elif 'date' in header_lower:
                widths.append(22 if num_cols < 14 else 18)
            elif 'status' in header_lower:
                widths.append(18 if num_cols < 14 else 15)
            elif 'cost' in header_lower or 'value' in header_lower or 'price' in header_lower:
                widths.append(20 if num_cols < 14 else 16)
            elif '%' in header_lower or 'percent' in header_lower:
                widths.append(16 if num_cols < 14 else 14)
            elif 'count' in header_lower:
                widths.append(16 if num_cols < 14 else 14)
            elif 'description' in header_lower:
                widths.append(40 if num_cols < 14 else 28)
            elif 'remarks' in header_lower or 'notes' in header_lower:
                widths.append(35 if num_cols < 14 else 25)
            elif 'category' in header_lower:
                widths.append(28 if num_cols < 14 else 20)
            elif 'location' in header_lower:
                widths.append(22 if num_cols < 14 else 18)
            elif 'site' in header_lower:
                widths.append(20 if num_cols < 14 else 16)
            elif 'department' in header_lower:
                widths.append(28 if num_cols < 14 else 20)
            elif 'supplier' in header_lower:
                widths.append(25 if num_cols < 14 else 18)
            elif 'brand' in header_lower:
                widths.append(20 if num_cols < 14 else 16)
            elif 'model' in header_lower:
                widths.append(20 if num_cols < 14 else 16)
            elif 'sku' in header_lower:
                widths.append(20 if num_cols < 14 else 16)
            elif 'barcode' in header_lower:
                widths.append(22 if num_cols < 14 else 18)
            elif 'name' in header_lower:
                widths.append(30 if num_cols < 14 else 22)
            elif 'employee' in header_lower:
                widths.append(30 if num_cols < 14 else 22)
            elif 'metric' in header_lower:
                widths.append(35 if num_cols < 14 else 25)
            elif 'method' in header_lower:
                widths.append(25 if num_cols < 14 else 20)
            elif 'accumulated' in header_lower:
                widths.append(28 if num_cols < 14 else 22)
            elif 'remaining' in header_lower:
                widths.append(22 if num_cols < 14 else 18)
            elif 'lessee' in header_lower:
                widths.append(28 if num_cols < 14 else 22)
            else:
                # Default based on header length, but respect min/max for column count
                default_width = max(min_width, min(max_width, len(header) * 2 + 8))
                widths.append(default_width)
        
        # Scale to fit page width exactly
        total = sum(widths)
        if total != page_width:
            ratio = page_width / total
            widths = [max(min_width, int(w * ratio)) for w in widths]
            
            # Adjust for rounding errors - distribute remainder to description/name columns
            remaining = int(page_width) - sum(widths)
            if remaining > 0:
                # Add remaining pixels to wider columns
                for i, header in enumerate(headers):
                    if remaining <= 0:
                        break
                    header_lower = header.lower()
                    if 'description' in header_lower or 'name' in header_lower or 'remarks' in header_lower:
                        widths[i] += 1
                        remaining -= 1
            
        return widths


def generate_pdf_from_excel_data(
    excel_content: bytes,
    report_name: str,
    report_type: str
) -> Optional[bytes]:
    """
    Generate PDF from Excel data by parsing the Excel file and 
    recreating it as a formatted PDF.
    """
    if not PDF_AVAILABLE:
        logger.error("fpdf2 not available for PDF generation")
        return None
        
    try:
        from openpyxl import load_workbook
        
        # Load the Excel workbook from bytes
        wb = load_workbook(io.BytesIO(excel_content))
        
        # Create PDF
        pdf = ReportPDF(report_name, report_type)
        pdf.add_page()
        
        # Define simplified columns per report type (to fit well in A4 Landscape PDF)
        REPORT_TYPE_COLUMNS = {
            "assets": [
                "Asset Tag ID", "Description", "Category", "Status", 
                "Cost", "Location", "Site", "Department"
            ],
            "location": [
                "Asset Tag", "Description", "Location", "Site", 
                "Department", "Status", "Last Move"
            ],
            "checkout": [
                "Asset Tag ID", "Description", "Category", "SUB-CATEGORY",
                "Check-out Date", "Due date", "Return Date", "Department", "Cost", "Employee"
            ],
            "maintenance": [
                "Asset Tag", "Description", "Title", "Status",
                "Due Date", "Completed", "Cost", "Inventory Items"
            ],
            "audit": [
                "Asset Tag ID", "Category", "Sub-Category", "Audit Type",
                "Audited to Site", "Audited to Location", "Last Audit Date", "Audit By"
            ],
            "depreciation": [
                "Asset Tag ID", "Description", "Category", "Depreciation Method",
                "Original Cost", "Depreciable Cost", "Accumulated Depreciation", "Current Value", "Date Acquired"
            ],
            "lease": [
                "Asset Tag ID", "Description", "Category", "Lessee",
                "Lease Start", "Lease End", "Status", "Days Remaining", "Asset Cost"
            ],
            "reservation": [
                "Asset Tag ID", "Description", "Category", "Reservation Type",
                "Reserved By", "Reservation Date", "Status", "Days Until/From", "Asset Cost"
            ],
            "transaction": [
                "Transaction Type", "Asset Tag ID", "Description", "Category",
                "Date", "Action By", "Details", "Location", "Asset Cost"
            ],
        }
        
        # Get simplified columns for this report type (default to assets)
        report_type_lower = report_type.lower()
        SIMPLIFIED_COLUMNS = REPORT_TYPE_COLUMNS.get(report_type_lower, REPORT_TYPE_COLUMNS["assets"])
        
        # Sheets that should use simplified columns (detail lists)
        DETAIL_SHEET_NAMES = [
            "Asset List", "Assets", "Asset Details", "Details",
            "Checkout List", "Maintenance List", "Audit List",
            "Lease List", "Reservation List", "Transaction List",
            "Location Assets", "Depreciation List"
        ]
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Add section for this sheet
            pdf.add_section_title(sheet_name)
            
            # Get all rows
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                pdf.set_font('Helvetica', 'I', 10)
                pdf.cell(0, 6, 'No data available', new_x='LMARGIN', new_y='NEXT')
                pdf.ln(5)
                continue
                
            # First row is header
            original_headers = [str(h) if h else '' for h in rows[0]]
            original_data_rows = [[str(c) if c else '' for c in row] for row in rows[1:]]
            
            # Check if this is a detail list sheet that should use simplified columns
            is_detail_sheet = any(
                name.lower() in sheet_name.lower() 
                for name in DETAIL_SHEET_NAMES
            )
            
            # Also check if headers contain many columns (detail list indicator)
            has_many_columns = (
                len(original_headers) > 10 and 
                any('asset' in h.lower() or 'tag' in h.lower() or 'id' in h.lower() for h in original_headers)
            )
            
            if is_detail_sheet or has_many_columns:
                # Use simplified columns for detail list
                # Map original headers to simplified ones
                header_indices = []
                simplified_headers = []
                
                for simplified_col in SIMPLIFIED_COLUMNS:
                    # Find matching column in original headers (case-insensitive, partial match)
                    for idx, orig_header in enumerate(original_headers):
                        # Normalize for comparison
                        orig_lower = orig_header.lower().replace(' ', '').replace('_', '')
                        simp_lower = simplified_col.lower().replace(' ', '').replace('_', '')
                        
                        if simp_lower in orig_lower or orig_lower in simp_lower:
                            header_indices.append(idx)
                            simplified_headers.append(simplified_col)
                            break
                
                if simplified_headers and header_indices:
                    # Extract only the simplified columns from data rows
                    simplified_data_rows = []
                    for row in original_data_rows:
                        simplified_row = []
                        for idx in header_indices:
                            if idx < len(row):
                                value = row[idx]
                                # Truncate long descriptions
                                if idx == header_indices[1] if len(header_indices) > 1 else -1:  # Description column
                                    value = value[:50] + '...' if len(value) > 50 else value
                                simplified_row.append(value)
                            else:
                                simplified_row.append('')
                        simplified_data_rows.append(simplified_row)
                    
                    if simplified_headers and simplified_data_rows:
                        pdf.add_table(simplified_headers, simplified_data_rows)
                    elif simplified_headers:
                        pdf.set_font('Helvetica', 'I', 10)
                        pdf.cell(0, 6, 'No records found', new_x='LMARGIN', new_y='NEXT')
                else:
                    # Fallback to original if mapping failed
                    if original_headers and original_data_rows:
                        pdf.add_table(original_headers, original_data_rows)
            else:
                # Non-asset-list sheets (Summary, By Status, By Category, etc.) - use all columns
                if original_headers and original_data_rows:
                    pdf.add_table(original_headers, original_data_rows)
                elif original_headers and not original_data_rows:
                    pdf.set_font('Helvetica', 'I', 10)
                    pdf.cell(0, 6, 'No records found', new_x='LMARGIN', new_y='NEXT')
                
            pdf.ln(10)
        
        # Return PDF content
        return bytes(pdf.output())
        
    except Exception as e:
        logger.error(f"Error generating PDF: {e}", exc_info=True)
        return None


def generate_simple_pdf(
    report_name: str,
    report_type: str,
    summary_data: Optional[Dict[str, Any]] = None,
    table_data: Optional[List[Dict[str, Any]]] = None
) -> Optional[bytes]:
    """
    Generate a simple PDF with summary and table data.
    Used when we have direct data instead of Excel content.
    """
    if not PDF_AVAILABLE:
        logger.error("fpdf2 not available for PDF generation")
        return None
        
    try:
        pdf = ReportPDF(report_name, report_type)
        pdf.add_page()
        
        # Add summary section if provided
        if summary_data:
            pdf.add_section_title('Summary')
            for key, value in summary_data.items():
                pdf.add_summary_row(key, str(value))
            pdf.ln(10)
        
        # Add table section if provided
        if table_data and len(table_data) > 0:
            pdf.add_section_title('Details')
            headers = list(table_data[0].keys())
            rows = [[str(row.get(h, '')) for h in headers] for row in table_data]
            pdf.add_table(headers, rows)
            
        return bytes(pdf.output())
        
    except Exception as e:
        logger.error(f"Error generating simple PDF: {e}", exc_info=True)
        return None


def is_pdf_available() -> bool:
    """Check if PDF generation is available"""
    return PDF_AVAILABLE

